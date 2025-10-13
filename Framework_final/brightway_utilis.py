import win32com.client as win32
from openpyxl import load_workbook
import pandas as pd
import bw2analyzer as ba
import bw2calc as bc
import bw2data as bd
import bw2io as bi
from rapidfuzz import process, fuzz

def init_excel(filepath = r"C:\Users\apuscas\Python\Environment\GitHubRep\Coding_AspenPlus-Brightway-LCA-Platform\E2DT2025\Framework_final\Inventory_Scenario_2_raw_file.xlsx"):
    global excel, wb, sheet
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = True
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(filepath, ReadOnly=True)
    sheet = wb.Sheets('Scenario_1_base case')

# Utility function to extract rows into DataFrame
def read_excel_row(skiprow, usecols, nrows=1, names=None):
        global sheet
        data = []
        for i in range(skiprow, skiprow + nrows):
            row = [(sheet.Cells(i+2, col+1)).Value for col in usecols]  
            data.append(row)
        df = pd.DataFrame(data, columns=names)
        return df

def close_excel():
    global excel, wb
    wb.Close(SaveChanges=False)
    excel.Quit()

def find_method(method_input):
    formatted_methods = [" | ".join(m) for m in bd.methods]
    match = process.extractOne(
        method_input.strip().lower(),
        [m.lower() for m in formatted_methods],
        scorer=fuzz.ratio,
        score_cutoff=60
    )
    if match:
        # Find the original tuple
        index = [m.lower() for m in formatted_methods].index(match[0])
        print(f"Loaded method: {list(bd.methods)[index]}")
        return (list(bd.methods)[index])
    else: 
        return print("Couldn't find method")

def clear_LCA_excel_output(name=str):
    print("Clearing " + name + ".xlsx")
    file_path = f"{name}.xlsx"
    sheet_name = "LCA_results"
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    for row in range(3, 51):
        for col in range(2, 5):
            ws.cell(row=row, column=col).value = None

    wb.save(file_path)

def calculate_LCA(functional_unit_activity_1, method_input):
    my_functional_unit, data_objs, _ = bd.prepare_lca_inputs(
        {functional_unit_activity_1: 1000},
        method=find_method(method_input)
    )
    my_lca = bc.LCA(demand=my_functional_unit, data_objs=data_objs)
    my_lca.lci()
    my_lca.lcia()
    return my_lca.score

def extract_res(result,i):
    with pd.ExcelWriter(r"LCA_results_rawfile.xlsx", mode="a" , engine="openpyxl", if_sheet_exists="overlay") as writer:
         result.to_excel(writer, header=False, sheet_name="LCA_results",index=False, startrow=i+1, startcol=1)

def search_row(search_value,filepath = r"C:\Users\apuscas\Python\Environment\GitHubRep\Coding_AspenPlus-Brightway-LCA-Platform\E2DT2025\Framework_final\Inventory_Scenario_2_raw_file.xlsx", column="B", sheet_name="Scenario_1_base case"):
    # Load the workbook in read-only mode
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]

    # Iterate through the given column
    for cell in ws[column]:
        if cell.value == search_value:
            return cell.row  # Return the row number if found

    return None  # If not found


if __name__=="__main__":
    row_number = search_row(column="B", search_value="Process Parameters")
    print(f"Found at row: {row_number}" if row_number else "Not found")
    row_number = search_row(column="B", search_value="Material Flow (Foreground System)")
    print(f"Found at row: {row_number}" if row_number else "Not found")
    row_number = search_row(column="B", search_value="Natural Resources (Background System)")
    print(f"Found at row: {row_number}" if row_number else "Not found")
    row_number = search_row(column="B", search_value="Energy Flow (Background)")
    print(f"Found at row: {row_number}" if row_number else "Not found")
    row_number = search_row(column="B", search_value="Infrastructure")
    print(f"Found at row: {row_number}" if row_number else "Not found")
